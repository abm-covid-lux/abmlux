#!/use/bin/env python3

#DensityModel

#This file generates a population density matrix using data collected by the GEOSTAT initiative. Note
#that Luxembourg measures 57 km x 82 km.

import numpy as np
from openpyxl import load_workbook
from tqdm import tqdm
import math

# Config
INPUT_FILENAME = 'Data/JRC.xlsx'
OUTPUT_FILENAME = 'Density_Map/Density_Map.csv'
LUXEMBOURG_SIZE_KM = (57, 82)   # width, height

# Load workbook
print(f"Loading input data from {INPUT_FILENAME}...")
wgtworkbook = load_workbook(filename=INPUT_FILENAME)
wgtsheet = wgtworkbook.active

# What are these parameters?
#xmin = 4014, xmax = 4070, xdifference = 56, ymin = 2934, ymax = 3015, ydifference = 81

D = [[0 for x in range(LUXEMBOURG_SIZE_KM[0])] for y in range(LUXEMBOURG_SIZE_KM[1])]

# TODO: Where are the numbers 2, 81809 coming from?
print(f"Calculating Density")
for s in tqdm(range(2,81809)):
    if (wgtsheet.cell(row=s, column=3).value == "LU"):
        x = int(wgtsheet.cell(row=s, column=2).value[9:13]) - 4014  # TODO: Why 4014?
        y = int(wgtsheet.cell(row=s, column=2).value[4:8]) - 2934   # TODO: Why 2934?
        D[y][x] = int(wgtsheet.cell(row=s, column=1).value)

print(f"Saving output to {OUTPUT_FILENAME}...")
np.savetxt(OUTPUT_FILENAME, D, fmt='%i', delimiter=',')
print('Done.') 
