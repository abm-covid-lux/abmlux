#!/usr/bin/env python3

#Markov Model

#This file constructs initial distributions and transition matrices using data from the Luxembourgish
#Time Use Survey (TUS). Each respondent to the TUS provided two diaries of activities, one for a week
#day and one for a weekend day, together with their age and the times at which these activities began
#. Associated to each respondent there is also a statistical weight and an identification number. Act
#ivities are denoted by a numeric code and therefore a daily routine consists of a sequence of such n
#umbers. Activities are grouped together according to the location in which the activity occurs. Dail
#y routines are then concatenated to produce, for each respondent, a typical weekly routine. The week
#begins on Sunday and ends on Saturday. Three sets of transition matrices and initial distributions a
#re then constructed from these routines, for children, adults and retired individuals, respectively.
#In particular, for each age group a transition matrix is constructed for each 10 minute interval of 
#the week, of which there are 7*144 in total.

import numpy as np
import math
from tqdm import tqdm
from openpyxl import load_workbook

INPUT_FILENAME = 'Data/TUS_Processed_2.xlsx'

print(f"Loading input data from {INPUT_FILENAME}...")
actworkbook = load_workbook(filename=INPUT_FILENAME)
actsheet = actworkbook.active

#First we count the number of dairies in the data:

samplesize = 0
print("Calculating sample size...")
for i in tqdm(range(2,97853)):
    if(actsheet.cell(row=i, column=4).value != actsheet.cell(row=i+1, column=4).value):
        samplesize = samplesize + 1

print('Generating daily routines...')

class Diary_Day:
  def __init__(self, identity, age, day, weight, daily_routine):
    self.identity = identity
    self.age = age
    self.day = day
    self.weight = weight
    self.daily_routine = daily_routine

D = [[0, 0, 0, 0, 0, []] for s in range(samplesize)]
for s in tqdm(range(samplesize)):
    D[s] = Diary_Day(0,0,0,0,[])

#As mentioned above, activities are numerically coded. The file FormatsTUS displays the codification 
#of primary and secondary activities used by the TUS. The primary activities will be grouped together
#according to datamap_primary while the secondary activities will be grouped together according to da
#tamap_secondary. The latter is used if and only if a respondent recorded 'Other specified location' 
#as the primary activity, in which case referal to the secondary activity is necessary. Activities ar
#e consquently recoded as numbers in the set {0,...,13}, as described in the file FormatActivities.

# FIXME: all the magic numbers in here!
def datamap_primary(x):
    if (x == 1):
        return 0
    if (x == 2):
        return 1
    if (x == 3):
        return 2
    if (x == 4):
        return 13
    if (x == 5):
        return 3
    if (x in {6,9,10,11,12,14}):
        return 4
    if (x == 13):
        return 5
    if (x in {15,16,17,18,19}):
        return 6

# FIXME: all the magic numbers in here!
def datamap_secondary(x):
    if (x in {11,12,13,14,21,22,23,31,34,35,39,115,121,213,214,221,222,231,239,311,314,315,324,325,326,327,328,329,333,346,347,349,351,353,354,356,363,364,371,381,382,383,384,386,391,393,419,421,422,431,432,433,434,511,512,522,551,711,713,714,719,733,734,737,739,744,745,746,747,749,811,812,813,819,821,829,839}):
        return 0
    if (x == 111):
        return 1
    if (x in {232,233}):
        return 2
    if (x in {395,426,429,515,521,523,524}):
        return 13
    if (x in {546,547}):
        return 3
    if (x in {343,345,411,412,413,415,424,439,525,539,541,542,543,545,548,549,612,613,614,617,618,619,629}):
        return 4
    if (x in {361,362,366,367,368,369,425}):
        return 7
    if (x == 365):
        return 8
    if (x == 435):
        return 9
    if (x in {544,615,616}):
        return 10
    if (x in {531,532,533}):
        return 11
    if (x == 534):
        return 12

#The following piece of code constructs the daily routines. This takes into account the fact that dia
#ries appearing in the TUS do not all start at the same time and do not all run for 24 hours. This is
#achieved by first extending the duration of the last activity, to create a 24 hour routine, after wh
#ich the piece of the routine extending beyong the end of the day is repositioned to the start of the
#day. This way, all routines cover a 24 hour period running from midnight to midnight.

s = 0
starttime = int(actsheet.cell(row=2, column=10).value)
for i in tqdm(range(2,97853)):     # FIXME: magic numbers
    time_now = actsheet.cell(row=i, column=10).value
    time_next = actsheet.cell(row=i+1, column=10).value
    if(actsheet.cell(row=i, column=4).value != actsheet.cell(row=i+1, column=4).value):
        D[s].identity = actsheet.cell(row=i, column=1).value
        D[s].age = actsheet.cell(row=i, column=3).value
        D[s].day = actsheet.cell(row=i, column=6).value
        D[s].weight = actsheet.cell(row=i, column=2).value
        time_next = 144 + starttime
        for j in range(time_next - time_now):
            if (actsheet.cell(row=i, column=9).value != 7):
                D[s].daily_routine.append(datamap_primary(actsheet.cell(row=i, column=9).value))
            if (actsheet.cell(row=i, column=9).value == 7):
                D[s].daily_routine.append(datamap_secondary(actsheet.cell(row=i, column=8).value))
        D[s].daily_routine = D[s].daily_routine[144-starttime:144] + D[s].daily_routine[0:144-starttime]
        s = s + 1
        starttime = int(actsheet.cell(row=i+1, column=10).value)
    else:
        for j in range(time_next - time_now):
            if (actsheet.cell(row=i, column=9).value != 7):
                D[s].daily_routine.append(datamap_primary(actsheet.cell(row=i, column=9).value))
            if (actsheet.cell(row=i, column=9).value == 7):
                D[s].daily_routine.append(datamap_secondary(actsheet.cell(row=i, column=8).value))

#For each respondent there are now two daily routines; one for a week day and one for a weekend day. 
#Copies of these routines are now concatenated so as to produce weekly routines, starting on Sunday.

print('Generating weekly routines...')
class Diary_Week:
  def __init__(self, identity, age, weight, weekly_routine):
    self.identity = identity
    self.age = age
    self.weight = weight
    self.weekly_routine = weekly_routine
 
numberofindividuals = int(samplesize/2)

W = [[0, 0, 0, 0, []]  for s in range(numberofindividuals)]

#The format of the TUS data is such that each pair of daily diaries are listed consecutively:

for s in tqdm(range(numberofindividuals)):
    if (D[2*s].day in {1,7}):
        weekend = D[2*s].daily_routine
        weekday = D[(2*s)+1].daily_routine
    else:
        weekend = D[(2*s)+1].daily_routine
        weekday = D[2*s].daily_routine
    W[s] = Diary_Week(D[2*s].identity,D[2*s].age,D[2*s].weight,weekend + weekday + weekday +  weekday + weekday + weekday + weekend)

#Now the statistical weights are used to construct the intial distributions and transition matrices:
print('Generating weighted initial distributions...')

Init_child = [0 for i in range(14)]
Init_adult = [0 for i in range(14)]
Init_retired = [0 for i in range(14)]

for s in tqdm(range(numberofindividuals)):
    if (W[s].age in range(10,18)):
        Init_child[W[s].weekly_routine[0]] = Init_child[W[s].weekly_routine[0]] + W[s].weight
    if (W[s].age in range(19,65)):
        Init_adult[W[s].weekly_routine[0]] = Init_adult[W[s].weekly_routine[0]] + W[s].weight
    if (W[s].age in range(65,76)):
        Init_retired[W[s].weekly_routine[0]] = Init_retired[W[s].weekly_routine[0]] + W[s].weight

np.savetxt('Initial_Distributions/Init_child.csv', Init_child, fmt='%i', delimiter=',')
np.savetxt('Initial_Distributions/Init_adult.csv', Init_adult, fmt='%i', delimiter=',')
np.savetxt('Initial_Distributions/Init_retired.csv', Init_retired, fmt='%i', delimiter=',')

print('Generating weighted transition matrices...')
Trans_child = [[[0 for i in range(14)] for j in range(14)] for t in range(7*144)]
Trans_adult = [[[0 for i in range(14)] for j in range(14)] for t in range(7*144)]
Trans_retired = [[[0 for i in range(14)] for j in range(14)] for t in range(7*144)]

#The transition matrices are now constructed, for all but the final ten minute interval:

for t in tqdm(range((7*144)-1)):
    for s in range(numberofindividuals):
        if (W[s].age in range(10,18)):
            Trans_child[t][W[s].weekly_routine[t]][W[s].weekly_routine[t+1]] = Trans_child[t][W[s].weekly_routine[t]][W[s].weekly_routine[t+1]] + W[s].weight
        if (W[s].age in range(19,65)):
            Trans_adult[t][W[s].weekly_routine[t]][W[s].weekly_routine[t+1]] = Trans_adult[t][W[s].weekly_routine[t]][W[s].weekly_routine[t+1]] + W[s].weight
        if (W[s].age in range(65,76)):
            Trans_retired[t][W[s].weekly_routine[t]][W[s].weekly_routine[t+1]] = Trans_retired[t][W[s].weekly_routine[t]][W[s].weekly_routine[t+1]] + W[s].weight
    np.savetxt('Transition_Matrices/Trans_child_' + str(t) + '.csv', Trans_child[t], fmt='%i', delimiter=',')
    np.savetxt('Transition_Matrices/Trans_adult_' + str(t) + '.csv', Trans_adult[t], fmt='%i', delimiter=',')
    np.savetxt('Transition_Matrices/Trans_retired_' + str(t) + '.csv', Trans_retired[t], fmt='%i', delimiter=',')

#The final transition matrix is constructed by looping the weekly routine back onto itself:
    
for s in tqdm(range(numberofindividuals)):
    if (W[s].age in range(10,18)):
        Trans_child[(7*144)-1][W[s].weekly_routine[(7*144)-1]][W[s].weekly_routine[0]] = Trans_child[(7*144)-1][W[s].weekly_routine[(7*144)-1]][W[s].weekly_routine[0]] + W[s].weight
    if (W[s].age in range(19,65)):
        Trans_adult[(7*144)-1][W[s].weekly_routine[(7*144)-1]][W[s].weekly_routine[0]] = Trans_adult[(7*144)-1][W[s].weekly_routine[(7*144)-1]][W[s].weekly_routine[0]] + W[s].weight
    if (W[s].age in range(65,76)):
        Trans_retired[(7*144)-1][W[s].weekly_routine[(7*144)-1]][W[s].weekly_routine[0]] = Trans_retired[(7*144)-1][W[s].weekly_routine[(7*144)-1]][W[s].weekly_routine[0]] + W[s].weight
np.savetxt('Transition_Matrices/Trans_child_' + str((7*144)-1) + '.csv', Trans_child[t], fmt='%i', delimiter=',')
np.savetxt('Transition_Matrices/Trans_adult_' + str((7*144)-1) + '.csv', Trans_adult[t], fmt='%i', delimiter=',')
np.savetxt('Transition_Matrices/Trans_retired_' + str((7*144)-1) + '.csv', Trans_retired[t], fmt='%i', delimiter=',')

print('Done.')
