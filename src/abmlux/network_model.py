"""This file procedurally generates the agents and locations."""

import math
import copy
from collections import defaultdict
import logging

from tqdm import tqdm
from scipy.spatial import KDTree
from openpyxl import load_workbook

from .random_tools import (multinoulli, multinoulli_dict, random_choices, random_sample,
                           random_choice, random_shuffle, random_randrange_interval)
from .agent import Agent, AgentType, POPULATION_SLICES
from .location import Location, WGS84_to_ETRS89
from .activity_manager import ActivityManager
from .network import Network
from .sim_time import SimClock

log = logging.getLogger('network_model')

def create_locations(network, density_map, config):
    """Create a number of Location objects within the network, according to the density map
    given and the distributions defined in the config. Locations with a non-deterministic count are
    created later."""

    log.debug('Initializing locations...')

    location_counts         = config['deterministic_location_counts']
    pop_by_border_countries = config['border_countries_pop']
    border_country_coord    = config['border_country_coord']
    pop_by_age              = config['age_distribution']

    # Adjust location counts by the ratio of the simulation size and real population size
    location_counts = {typ: math.ceil((config['n'] / sum(pop_by_age)) * location_counts[typ])
                       for typ, x in location_counts.items()}
    location_counts['Outdoor'] = 1
    log.debug("Location count by type: %s", location_counts)
    # Create locations for each type, of the amounts requested
    log.info("Constructing locations...")
    for ltype, lcount in location_counts.items():
        for _ in range(lcount):
            new_coord = density_map.sample_coord()
            new_location = Location(ltype, new_coord)
            network.add_location(new_location)
    # Create locations of each border country
    for country in pop_by_border_countries:
        coord = WGS84_to_ETRS89(border_country_coord[country][1], border_country_coord[country][0])
        new_country = Location(country, coord)
        network.add_location(new_country)

def create_agents(prng, network, config):
    """Create a number of Agent objects within the network, according to the distributions
    specified in the configuration object provided."""

    log.debug('Initializing agents...')

    pop_by_age = config['age_distribution']

    # How many agents per agent type
    pop_normalised    = [x / sum(pop_by_age) for x in pop_by_age]
    pop_by_agent_type = {atype: math.ceil(config['n'] * sum(pop_normalised[slce]))
                         for atype, slce in POPULATION_SLICES.items()}
    # Add an appropriate number of cross border workers as adults
    pop_by_border_country = config['border_countries_pop']
    total = sum(pop_by_border_country.values())
    pop_by_agent_type[AgentType.ADULT] += math.ceil(total * config['n']/sum(pop_by_age))
    log.debug("Agent count by type: %s", pop_by_agent_type)
    # The total numbers of children, adults and retired individuals are fixed deterministically,
    # while the exact age of individuals within each group is determined randomly
    log.info("Constructing agents...")
    for atype, slce in POPULATION_SLICES.items():
        log.info(" - %s...", atype.name)
        for _ in tqdm(range(pop_by_agent_type[atype])):
            # Sample a point from the age distribution within this slice
            age = (slce.start or 0) + multinoulli(prng, pop_by_age[slce])
            new_agent = Agent(atype, age)
            network.add_agent(new_agent)

def make_house_profile_dictionary(config):
    """Creates a probability distribution across household profiles."""

    log.debug("Making housing dictionary...")

    hshld_dst_c = config['household_distribution_children']
    hshld_dst_r = config['household_distribution_retired']

    max_house_size = len(hshld_dst_r[0]) - 1
    if max_house_size != len(hshld_dst_c[0]) - 1:
        raise Exception("Distributions of children and retired are in conflict: max house size")
    total_houses = sum([sum(x) for x in zip(*hshld_dst_r)])
    if total_houses != sum([sum(x) for x in zip(*hshld_dst_c)]):
        raise Exception("Distributions of children and retired are in conflict: total houses")
    # Each key in the following dictionary, house_profiles, will be a triple. The entries in each
    # triple will correspond to numbers of children, adults and retired, respectively. The
    # corresponding value indicates the probility of that triple occuring as household profile.
    # The sum of the entries in a given triple is bounded by the maximum house size. All possible
    # keys are generated that satisfy this bound.
    #
    # To generate the probabilties, it is assumed that, conditional on the house size being n and
    # the number of retired residents in a house being r, the number of children c follows the
    # distribution given by normalizing the first n-r entries in the n-th column of the matrix
    # hshld_dst_c.
    house_profiles = {}
    for house_size in range(1, max_house_size + 1):
        for num_children in range(house_size + 1):
            for num_retired in range(house_size + 1 - num_children):
                num_adult = house_size - num_children - num_retired
                weight = sum(tuple(zip(*hshld_dst_c))[house_size][0:house_size + 1 - num_retired])
                prob = hshld_dst_c[num_children][house_size]\
                       * hshld_dst_r[num_retired][house_size] / (total_houses*weight)
                house_profiles[(num_children, num_adult, num_retired)] = prob

    return house_profiles

def assign_homes(prng, network, density_map, config, activity_manager, house_location_type,
                 home_activity_type, carehome_type):
    """Assigns homes to agents."""

    log.info("Creating and populating homes...")

    unassigned_children = copy.copy(network.agents_by_type[AgentType.CHILD])
    unassigned_adults   = copy.copy(network.agents_by_type[AgentType.ADULT])
    unassigned_retired  = copy.copy(network.agents_by_type[AgentType.RETIRED])

    # ---- Populate Carehomes ----
    log.debug("Populating care homes...")
    # Number of residents per carehome
    num_retired = config['retired_per_carehome']
    carehomes = copy.copy(network.locations_for_types(carehome_type))
    occupancy_carehomes = {}
    for carehome in carehomes:
        # Take from front of lists
        carehome_residents = unassigned_retired[0:num_retired]
        del unassigned_retired[0:num_retired]
        # Assign agents to carehome
        occupancy_carehomes[carehome] = carehome_residents
        for resident in carehome_residents:
            resident.add_activity_location(activity_manager.as_int(home_activity_type), carehome)

    # ---- Populate Border Countries ----
    log.debug("Populating border countries...")
    # Cross border worker populations
    pop_by_border_country = config['border_countries_pop']
    total_pop_by_age = sum(config['age_distribution'])
    occupancy_border_countries = {}
    for border_country in pop_by_border_country:
        # Normalize cross border worker populations
        bc_n = math.ceil(config['n']*pop_by_border_country[border_country]/total_pop_by_age)
        # Take from front of lists
        border_country_workers = unassigned_adults[0:bc_n]
        del unassigned_adults[0:bc_n]
        country = network.locations_for_types(border_country)[0]
        occupancy_border_countries[country] = border_country_workers
        for worker in border_country_workers:
            worker.add_activity_location(activity_manager.as_int(home_activity_type), country)

    # ---- Populate Houses ----
    log.debug("Populating houses...")
    # Type distribution from which to sample
    house_types = make_house_profile_dictionary(config)
    occupancy_houses = {}
    while len(unassigned_children + unassigned_adults + unassigned_retired) > 0:
        # Create new house and add it to the network
        house_coord = density_map.sample_coord()
        new_house = Location(house_location_type, house_coord)
        network.add_location(new_house)
        # Generate household profile
        household_profile = multinoulli_dict(prng, house_types)
        num_children = min(household_profile[0], len(unassigned_children))
        num_adults   = min(household_profile[1], len(unassigned_adults))
        num_retired  = min(household_profile[2], len(unassigned_retired))
        # Take agents from front of lists
        children = unassigned_children[0:num_children]
        del unassigned_children[0:num_children]
        adults = unassigned_adults[0:num_adults]
        del unassigned_adults[0:num_adults]
        retired = unassigned_retired[0:num_retired]
        del unassigned_retired[0:num_retired]
        # Assign agents to new house
        occupancy_houses[new_house] = children + adults + retired
        for occupant in occupancy_houses[new_house]:
            occupant.add_activity_location(activity_manager.as_int(home_activity_type), new_house)

    return occupancy_houses, occupancy_carehomes, occupancy_border_countries

def do_activity_from_home(activity_manager, occupancy, activity_type):
    """Sets the activity location as the occupancy location, for all agents listed in an
    occupancy dictionary."""

    for location in occupancy:
        for agent in occupancy[location]:
            agent.add_activity_location(activity_manager.as_int(activity_type), location)

def make_distribution(config, motive, country_origin, country_destination,
                            number_of_bins, bin_width):
    """For given country  of origin, country of destination and motive, this creates a probability
    distribution over ranges of distances."""

    log.info("Generating distance distribution...")

    actworkbook = load_workbook(filename = config.filepath('trip_data_filepath'))
    actsheet    = actworkbook.active

    # In the following distribution, the probability assigned to a given range reflects the
    # probability that the length of a trip, between the input countries and with the given
    # motivation, falls within that range. Note that the units of bid_width are kilometers, and that
    # the distances recorded in the data refer to distance travelled by the respondent, not as the
    # crow flies.
    distance_distribution = {}
    for bin_num in range(number_of_bins):
        distance_distribution[range(bin_width*bin_num,bin_width*(bin_num+1))] = 0
    for sheet_row in tqdm(range(2,actsheet.max_row)):
        motive_sample = actsheet.cell(row=sheet_row, column=7).value
        country_origin_sample = actsheet.cell(row=sheet_row, column=9).value
        country_destination_sample = actsheet.cell(row=sheet_row, column=11).value
        # For each sample of the desired trip type, record the distance and add to the distribution
        if ([motive_sample,country_origin_sample,country_destination_sample]
            == [motive, country_origin, country_destination]):
            distance = actsheet.cell(row=sheet_row, column=12).value
            if isinstance(distance,(int,float)) and (distance < number_of_bins*bin_width):
                weight = actsheet.cell(row=sheet_row, column=15).value
                distance_distribution[range(int((distance//bin_width)*bin_width),
                          int(((distance//bin_width)+1)*bin_width))] += round(weight)
    # Normalize to obtain a probability distribution
    total_weight = sum(distance_distribution.values())
    for distribution_bin in distance_distribution:
        distance_distribution[distribution_bin] /= total_weight

    return distance_distribution

def road_distance(config, euclidean_distance_km):
    """Converts a Euclidean distance into a network distance."""

    alpha = config['alpha']
    beta  = config['beta']

    return (euclidean_distance_km * alpha) + beta

def euclidean_distance(coords1, coords2):
    """Calculates the Euclidean distance between two points."""

    return math.sqrt(((coords1[0]-coords2[0])**2) + ((coords1[1]-coords2[1])**2))

def get_weight(config, dist_km, distance_distribution):
    """Given a distance, in kilometers, and a distance_distribution, returns the probability weight
    associated to that distance by the distribution."""

    dist_length = sum([len(dist_bin) for dist_bin in list(distance_distribution.keys())])
    if int(road_distance(config, dist_km)) >= dist_length:
        return 0.0
    else:
        for distribution_bin in distance_distribution:
            if int(road_distance(config, dist_km)) in distribution_bin:
                return distance_distribution[distribution_bin]
                break

def make_work_profile_dictionary(prng, network, config):
    """Generates weights for working locations"""

    workforce_profile_distribution = config['workforce_profile_distribution']
    workforce_profile_uniform      = config['workforce_profile_uniform']
    profile_format                 = config['workforce_profile_distribution_format']

    # Weights reflect typical size of workforce in locations across different sectors
    workplace_weights = {}
    for location_type in workforce_profile_distribution:
        profile = workforce_profile_distribution[location_type]
        for location in network.locations_by_type[location_type]:
            interval = profile_format[multinoulli(prng, profile)]
            weight = random_randrange_interval(prng, interval[0],interval[1])
            workplace_weights[location] = weight
    for location_type in workforce_profile_uniform:
        weight = workforce_profile_uniform[location_type]
        for location in network.locations_by_type[location_type]:
            workplace_weights[location] = weight

    return workplace_weights

def assign_workplaces(prng, network, config, activity_manager, work_activity_type,
                      occupancy_houses, occupancy_carehomes, occupancy_border_countries):
    """Assign a place of work for each agent."""

    log.info("Assigning places of work...")

    bin_width           = config['bin_width']
    number_of_bins      = config['number_of_bins']
    sample_size         = config['location_sample_size']
    destination_country = config['destination_country']
    origin_country_dict = config['origin_country_dict']
    activity_dict       = config['activity_dict']

    # These determine the probability of an agent travelling a distance to work
    work_dist_dict = {}
    for country in origin_country_dict:
        work_dist_dict[country] = make_distribution(config, activity_dict[work_activity_type],
                                            origin_country_dict[country], destination_country,
                                            number_of_bins[country], bin_width[country])

    log.info("Generating workforce weights...")
    # These weights corrspond to the size of the workforce at each workplace
    workplace_weights = make_work_profile_dictionary(prng, network, config)
    log.info("Assigning workplaces to house occupants...")
    wrkplaces = network.locations_for_types(activity_manager.get_location_types(work_activity_type))
    for house in tqdm(occupancy_houses):
        # Here each house gets a sample from which occupants choose
        work_locations_sample = random_sample(prng, wrkplaces, k = min(sample_size, len(wrkplaces)))
        weights_for_house = {}
        for location in work_locations_sample:
            dist_m = euclidean_distance(house.coord, location.coord)
            dist_km = dist_m/1000
            weight = get_weight(config, dist_km, work_dist_dict['Luxembourg'])
            # For each location, the workforce weights and distance weights are multiplied
            weights_for_house[location] = workplace_weights[location] * weight
        for agent in occupancy_houses[house]:
            # A workplace is then chosen randomly from the sample, according to the weights
            workplace = multinoulli_dict(prng, weights_for_house)
            agent.add_activity_location(activity_manager.as_int(work_activity_type), workplace)
        weights_for_house.clear()

    log.info("Assigning workplaces to border country occupants...")
    for border_country in occupancy_border_countries:
        for agent in tqdm(occupancy_border_countries[border_country]):
            # Here each agent gets a sample from which to choose
            work_locations_sample = random_sample(prng, wrkplaces, k = min(sample_size, len(wrkplaces)))
            weights_for_agent = {}
            for location in work_locations_sample:
                dist_m = euclidean_distance(border_country.coord, location.coord)
                dist_km = dist_m/1000
                weight = get_weight(config, dist_km, work_dist_dict[border_country.typ])
                weights_for_agent[location] = workplace_weights[location] * weight
            workplace = multinoulli_dict(prng, weights_for_agent)
            agent.add_activity_location(activity_manager.as_int(work_activity_type), workplace)
            weights_for_agent.clear()

    log.debug("Assigning workplaces to carehome occupants...")
    do_activity_from_home(activity_manager, occupancy_carehomes, work_activity_type)

def assign_locations_by_distance(prng, network, config, activity_manager, activity_type,
                                 occupancy_houses, occupancy_carehomes, occupancy_border_countries):
    """Assign activities to agents by distance"""
    # For each individual, a number of distinct locations, not including the individual's own home,
    # are randomly selected so that the individual is able to visit them:
    log.info("Assigning locations by distance for activity: %s...", activity_type)

    bin_width      = config['bin_width']
    number_of_bins = config['number_of_bins']
    num_can_visit  = config['activity_locations_by_distance']
    sample_size    = config['location_sample_size']
    activity_dict  = config['activity_dict']

    vst_locs = network.locations_for_types(activity_manager.get_location_types(activity_type))
    # This determines the probability of an agent travelling a distance for a house visit
    dist_dict = make_distribution(config, activity_dict[activity_type], 'Luxembourg', 'Luxembourg',
                                  number_of_bins['Luxembourg'], bin_width['Luxembourg'])
    log.debug("Assigning locations to house occupants...")
    for house in tqdm(occupancy_houses):
        visit_locations_sample = random_sample(prng, vst_locs, k = min(sample_size, len(vst_locs)))
        weights_for_house = {}
        for location in visit_locations_sample:
            dist_m = euclidean_distance(house.coord, location.coord)
            dist_km = dist_m/1000
            weights_for_house[location] = get_weight(config, dist_km, dist_dict)
        for agent in occupancy_houses[house]:
            # Several houses are then chosen randomly from the sample, according to the weights
            locs = random_choices(prng, list(weights_for_house.keys()),
                                  list(weights_for_house.values()), num_can_visit[activity_type])
            # If the activity is visit and the agent's own home is chosen, then it is removed from
            # the list and the sample can therefore be of size num_can_visit['Visit']-1
            if (activity_type == 'Visit') and (house in locs):
                locs.remove(house)
            agent.add_activity_location(activity_manager.as_int(activity_type), locs)
        weights_for_house.clear()
    log.debug("Assigning locations to border country occupants...")
    do_activity_from_home(activity_manager, occupancy_border_countries, activity_type)
    log.debug("Assigning locations to carehome occupants...")
    do_activity_from_home(activity_manager, occupancy_carehomes, activity_type)

def assign_locations_by_random(prng, network, config, activity_manager, activity_type,
                               occupancy_houses, occupancy_carehomes, occupancy_border_countries):
    """For each agent, a number of distinct locations are randomly selected"""

    log.info("Assigning locations by random for activity: %s...", activity_type)

    num_can_visit = config['activity_locations_by_random']

    venues = network.locations_for_types(activity_manager.get_location_types(activity_type))
    log.debug("Assigning locations by random to house occupants...")
    for house in tqdm(occupancy_houses):
        for agent in occupancy_houses[house]:
            venues_sample = random_sample(prng, venues, k=min(len(venues), num_can_visit[activity_type]))
            agent.add_activity_location(activity_manager.as_int(activity_type), venues_sample)
    log.debug("Assigning locations by random to border country occupants...")
    do_activity_from_home(activity_manager, occupancy_border_countries, activity_type)
    log.debug("Assigning locations by random to carehome occupants...")
    do_activity_from_home(activity_manager, occupancy_carehomes, activity_type)

def assign_locations_by_proximity(prng, network, activity_manager, activity_type, occupancy_houses,
                                  occupancy_carehomes, occupancy_border_countries):
    """For the location type given, select nearby houses and assign all occupants to
    attend this location.  If the location is full, move to the next nearby location, etc."""

    log.info("Assigning proximate locations for activity: %s...", activity_type)

    # The following code assigns homes to locations in such a way that equal numbers of homes are
    # assigned to each location of a given type. For example, from the list of homes, a home is
    # randomly selected and assigned to the nearest school, unless that school has already been
    # assigned its share of homes, in which case the next nearest available school is assigned.
    # This creates local spatial structure while ensuring that no school, for example, is
    # assigned more homes than the other schools. This same procedure is also applied to medical
    # locations, places of worship and indoor sport:
    log.debug("Assigning proximate locations to house occupants...")
    log.debug("Finding people to perform activity: %s", activity_type)
    log.debug("Location types: %s", activity_manager.get_location_types(activity_type))

    # Ensure we have at least one location of this type
    locations = network.locations_for_types(activity_manager.get_location_types(activity_type))
    log.debug("Found %i available locations", len(locations))
    assert len(locations) > 0

    max_homes  = math.ceil(network.count('House') / len(locations))
    kdtree     = KDTree([l.coord for l in locations])
    num_houses = defaultdict(int)

    # Traverse houses in random order
    shuffled_houses = copy.copy(network.locations_by_type['House'])
    random_shuffle(prng, shuffled_houses)
    for house in tqdm(shuffled_houses):
        # Find the closest location and, if it's not full, assign every occupant to the location
        knn = 2
        closest_locations = []
        while len(closest_locations) == 0:
            if (knn/2) > len(locations):
                raise ValueError(f"Searching for more locations than exist for "
                                 f"types {activity_manager.get_location_types(activity_type)}.  "
                                 f"This normally indicates that all locations are full.")
            # Returns knn items, in order of nearness
            _, nearest_indices = kdtree.query(house.coord, knn)
            closest_locations = [locations[i] for i in nearest_indices if i < len(locations)]
            # Remove locations that have too many houses already
            closest_locations = [x for x in closest_locations if num_houses[x] < max_homes]
            knn *= 2
        closest_location = closest_locations[0]
        # Add all occupants of this house to the location
        num_houses[closest_location] += 1
        for occupant in occupancy_houses[house]:
            occupant.add_activity_location(activity_manager.as_int(activity_type), closest_location)
    log.debug("Assigning proximate locations to border country occupants...")
    do_activity_from_home(activity_manager, occupancy_border_countries, activity_type)
    log.debug("Assigning proximate locations to carehome occupants...")
    do_activity_from_home(activity_manager, occupancy_carehomes, activity_type)

def assign_outdoors(network, activity_manager, outdoor_activity_type, occupancy_houses,
                    occupancy_carehomes, occupancy_border_countries):
    """Ensure all residents except carehome residents are allowed to access the outdoors"""

    log.info("Assigning outdoor location...")

    outdrs = network.locations_for_types(activity_manager.get_location_types(outdoor_activity_type))
    #The outdoors is treated as a single environment, in which zero disease transmission will occur
    if len(outdrs) != 1:
        raise ValueError("More than one outdoor location found. Set outdoor count to 1.")
    outdrs_loc = outdrs[0]
    log.debug("Assigning outdoor location to house occupants...")
    for house in tqdm(occupancy_houses):
        for agent in occupancy_houses[house]:
            agent.add_activity_location(activity_manager.as_int(outdoor_activity_type), outdrs_loc)
    log.debug("Assigning outdoor location to border country occupants...")
    do_activity_from_home(activity_manager, occupancy_border_countries, outdoor_activity_type)
    log.debug("Assigning outdoor location to carehome occupants...")
    do_activity_from_home(activity_manager, occupancy_carehomes, outdoor_activity_type)

def assign_cars(network, activity_manager, car_activity_type, car_location_type, occupancy_houses,
                occupancy_carehomes, occupancy_border_countries):
    """Assign a car to each house. All occupants of a house use the same car."""

    log.info("Assigning car location for activity %s...", car_activity_type)

    log.debug("Assigning car to house occupants...")
    for house in tqdm(occupancy_houses):
        new_car = Location(car_location_type, house.coord)
        network.add_location(new_car)
        for agent in occupancy_houses[house]:
            agent.add_activity_location(activity_manager.as_int(car_activity_type), new_car)
    log.debug("Assigning car to border country occupants...")
    do_activity_from_home(activity_manager, occupancy_border_countries, car_activity_type)
    log.debug("Assigning car to carehome occupants...")
    do_activity_from_home(activity_manager, occupancy_carehomes, car_activity_type)

def build_network_model(prng, config, density_map):
    """Create agents and locations according to the population density map given"""

    activity_manager = ActivityManager(config['activities'])

    log.info("Creating network...")

    network = Network(density_map)
    create_agents(prng, network, config)
    create_locations(network, density_map, config)

    log.info("Assigning locations to agents...")

    # Assign homes
    occupancy_houses, occupancy_carehomes, occupancy_border_countries\
        = assign_homes(prng, network, density_map, config, activity_manager,
                       "House", "House", "Care Home")
    # Assign workplaces
    assign_workplaces(prng, network, config, activity_manager, "Work", occupancy_houses,
                      occupancy_carehomes, occupancy_border_countries)
    # Assignments of locations by distance
    for activity_type in config['activity_locations_by_distance']:
        assign_locations_by_distance(prng, network, config, activity_manager, activity_type,
                                     occupancy_houses, occupancy_carehomes,
                                     occupancy_border_countries)
    # Assignments of locations by random
    for activity_type in config['activity_locations_by_random']:
        assign_locations_by_random(prng, network, config, activity_manager, activity_type,
                                   occupancy_houses, occupancy_carehomes,
                                   occupancy_border_countries)
    # Assignments of locations by proximity
    for activity_type in config['activity_locations_by_proximity']:
        assign_locations_by_proximity(prng, network, activity_manager, activity_type,
                                      occupancy_houses, occupancy_carehomes,
                                      occupancy_border_countries)
    # Assignments of outdoors
    assign_outdoors(network, activity_manager, "Outdoor", occupancy_houses, occupancy_carehomes,
                    occupancy_border_countries)
    # Assignments of cars
    assign_cars(network, activity_manager, "Car", "Car", occupancy_houses, occupancy_carehomes,
                occupancy_border_countries)

    return network

def assign_activities(prng, config, network, activity_distributions):
    """Assign activities and locations to agents according to the distributions provided. This
    moreover generates the initial state."""

    clock = SimClock(config['tick_length_s'], config['simulation_length_days'], config['epoch'])
    log.debug("Loading initial state for simulation...")

    log.debug("Seeding initial activity states and locations...")
    for agent in network.agents:
        # Get distribution for this type at the starting time step
        distribution = activity_distributions[agent.agetyp][clock.epoch_week_offset]
        assert sum(distribution.values()) > 0
        new_activity      = multinoulli_dict(prng, distribution)
        allowed_locations = agent.locations_for_activity(new_activity)
        # Warning: No allowed locations found for agent {agent.inspect()} for activity new_activity
        assert len(allowed_locations) >= 0
        new_location = random_choice(prng, list(allowed_locations))
        # Do this activity in a random location
        agent.set_activity(new_activity, new_location)

    return network
