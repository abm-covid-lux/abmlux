#!/usr/bin/env python3

#Network Simulation

#This file plots the procedurally generated environment constructed in the file NetworkModel.

import numpy as np
from openpyxl import load_workbook
import math
import matplotlib.pyplot as plt

from agent import Agent

OUTPUT_MAP_FILENAME = "Results/Network_map.png"
FIGURE_SIZE         = (20, 20)


agentworkbook = load_workbook(filename='Agents/Agents.xlsx')
agentsheet = agentworkbook.active

N = agentsheet.max_row

P = [[0, 0] for i in range(N)]

for i in range(N):
    P[i] = Agent(agentsheet.cell(row=i+1, column=2).value,agentsheet.cell(row=i+1, column=3).value)

LocationListAgent = [ [ [] for j in range(14) ] for i in range(N) ]
    
for i in range(N):
    for j in range(14):        
        LocationListAgent[i][j] = list(map(int,agentsheet.cell(row=i+1, column=j+4).value.split(',')))
        
class Location:
  def __init__(self, typ, coord, who):
    self.typ = typ
    self.coord = coord
    self.who = who
    
locationworkbook = load_workbook(filename='Locations/Locations.xlsx')
locationsheet = locationworkbook.active



print(f"Rendering figure to {OUTPUT_MAP_FILENAME}...")
M = locationsheet.max_row
L = [ [ 0, [0,0] ] for i in range(M) ]
plt.figure(figsize=FIGURE_SIZE)
for j in range(M):
    L[j] = Location(locationsheet.cell(row=j+1, column=2).value,list(map(int,locationsheet.cell(row=j+1, column=3).value.split(','))),[])
    if (L[j].typ == 0):
        plt.plot(L[j].coord[0],L[j].coord[1], marker = ".", color='green', label="House")
    if (L[j].typ == 1):
        plt.plot(L[j].coord[0],L[j].coord[1], marker = ".", color='red', label="Other Work")
    if (L[j].typ == 2):
        plt.plot(L[j].coord[0],L[j].coord[1], marker = "s", color='blue', label="Schools")
    if (L[j].typ == 3):
        plt.plot(L[j].coord[0],L[j].coord[1], marker = "o", color='purple', label="Restaurants")
    if (L[j].typ == 7):
        plt.plot(L[j].coord[0],L[j].coord[1], marker = "s", color='cyan', label="Shops")
    if (L[j].typ == 8):
        plt.plot(L[j].coord[0],L[j].coord[1], marker = "P", color='pink', label="Medical")
    if (L[j].typ == 9):
        plt.plot(L[j].coord[0],L[j].coord[1], marker = "+", color='black', label="Place of Worship")
    if (L[j].typ == 10):
        plt.plot(L[j].coord[0],L[j].coord[1], marker = "$s$", color='black', label="Indoor Sport")
    if (L[j].typ == 11):
        plt.plot(L[j].coord[0],L[j].coord[1], marker = "$c$", color='black', label="Cinema or Theatre")
    if (L[j].typ == 12):
        plt.plot(L[j].coord[0],L[j].coord[1], marker = "$m$", color='black', label="Museum or Zoo")

plt.savefig(OUTPUT_MAP_FILENAME, dpi=100)
print('Done.')
